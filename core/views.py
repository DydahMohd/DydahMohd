import os
from datetime import datetime, timedelta
from io import BytesIO, StringIO
import csv
import math
from xml.sax.saxutils import escape

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.views import LoginView
from django.core.mail import send_mail
from django.db.models import Count, Q, F
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import CreateView, DetailView, FormView, ListView, TemplateView, UpdateView

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph, Image
from reportlab.pdfgen.canvas import Canvas
from reportlab.lib.units import inch

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
except Exception:
    plt = None


from .forms import (
    CustomAuthenticationForm, 
    CustomUserCreationForm, 
    TicketAttachmentForm, 
    TicketCommentForm, 
    TicketForm, 
    TicketReopenForm, 
    TicketUpdateForm, 
    UserRoleForm,
    DeviceForm,
    KnowledgeBaseForm
)
from .models import (
    AuditLog,
    CustomUser,
    Ticket,
    TicketPriority,
    TicketStatus,
    UserRole,
    WingChoices,
    FloorChoices,
    Device,
    DeviceCondition,
    DeviceStatus,
    TicketComment,
    KnowledgeBaseArticle,
    Material,
    MaterialRequest,
    MaterialRequestStatus,
)
from .forms import MaterialRequestForm, MaterialForm
from .utils import (
    get_stale_tickets, 
    send_stale_ticket_notification_emails, 
    send_overdue_ticket_notification_emails,
    predict_ticket_category,
)




def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '')


def get_report_ticket_queryset(user):
    if user.is_admin:
        return Ticket.objects.select_related('reported_by', 'assigned_to').all()
    return Ticket.objects.select_related('reported_by', 'assigned_to').filter(
        Q(reported_by=user) | Q(assigned_to=user)
    )


def user_can_access_ticket(user, ticket):
    if user.is_admin:
        return True
    if user.is_technician:
        return ticket.assigned_to == user or ticket.status == TicketStatus.OPEN or ticket.reported_by == user
    return ticket.reported_by == user


def user_can_modify_ticket(user, ticket):
    if user.is_admin or user.is_technician:
        return True
    return ticket.status not in (TicketStatus.RESOLVED, TicketStatus.CLOSED)


def send_notification_email(subject, message, recipients):
    if not recipients:
        return
    send_mail(
        subject,
        message,
        settings.DEFAULT_FROM_EMAIL,
        recipients,
        fail_silently=True,
    )


def apply_ticket_report_filters(queryset, params):
    status = params.get('status')
    priority = params.get('priority')
    category = params.get('category')
    date_from = params.get('date_from')
    date_to = params.get('date_to')

    if status:
        queryset = queryset.filter(status=status)
    if priority:
        queryset = queryset.filter(priority=priority)
    if category:
        queryset = queryset.filter(category__icontains=category)

    if date_from:
        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            queryset = queryset.filter(created_at__date__gte=start_date)
        except ValueError:
            pass

    if date_to:
        try:
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            queryset = queryset.filter(created_at__date__lte=end_date)
        except ValueError:
            pass

    return queryset


def apply_audit_filters(queryset, params):
    action_type = params.get('action_type')
    user_filter = params.get('user')
    ticket_filter = params.get('ticket')
    date_from = params.get('date_from')
    date_to = params.get('date_to')

    if action_type:
        queryset = queryset.filter(action_type__icontains=action_type)
    if user_filter:
        queryset = queryset.filter(
            Q(user__username__icontains=user_filter) |
            Q(user__first_name__icontains=user_filter) |
            Q(user__last_name__icontains=user_filter)
        )
    if ticket_filter:
        queryset = queryset.filter(
            Q(ticket__title__icontains=ticket_filter) |
            Q(ticket__id__icontains=ticket_filter)
        )

    if date_from:
        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            queryset = queryset.filter(action_time__date__gte=start_date)
        except ValueError:
            pass

    if date_to:
        try:
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            queryset = queryset.filter(action_time__date__lte=end_date)
        except ValueError:
            pass

    return queryset


class UserLoginView(LoginView):
    template_name = 'core/login.html'
    authentication_form = CustomAuthenticationForm


class UserLogoutView(View):
    """
    Handles user logout. Calls logout(request) to clear the session and 
    records the action in the AuditLog.
    """
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            method_label = " (GET)" if request.method == "GET" else ""
            AuditLog.objects.create(
                user=request.user,
                action_type='user_logged_out',
                description=f"User '{request.user.username}' logged out{method_label}.",
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
            )
        logout(request)
        return redirect('login')


class RegisterView(FormView):
    template_name = 'core/register.html'
    form_class = CustomUserCreationForm
    success_url = reverse_lazy('login')

    def form_valid(self, form):
        form.save()
        return super().form_valid(form)


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'core/dashboard.html'
    login_url = reverse_lazy('login')

    def get_context_data(self, **kwargs):
        context = kwargs.copy()
        context['view'] = self

        user = self.request.user
        if user.is_admin:
            queryset = Ticket.objects.all()
        elif user.is_technician:
            queryset = Ticket.objects.filter(
                Q(assigned_to=user) | Q(status=TicketStatus.OPEN) | Q(reported_by=user)
            )
        else:
            queryset = Ticket.objects.filter(reported_by=user)

        context['user_role'] = user.get_role_display() if hasattr(user, 'get_role_display') else getattr(user, 'role', '')
        context['is_admin'] = getattr(user, 'is_admin', False)
        context['is_technician'] = getattr(user, 'is_technician', False)
        context['is_property_manager'] = getattr(user, 'is_property_manager', False)
        context['total_tickets'] = queryset.count()
        context['open_tickets'] = queryset.filter(status=TicketStatus.OPEN).count()
        context['in_progress_tickets'] = queryset.filter(status=TicketStatus.IN_PROGRESS).count()
        context['resolved_tickets'] = queryset.filter(status=TicketStatus.RESOLVED).count()
        context['closed_tickets'] = queryset.filter(status=TicketStatus.CLOSED).count()
        
        active_tickets = queryset.filter(status__in=[TicketStatus.OPEN, TicketStatus.IN_PROGRESS])
        context['high_priority_tickets'] = active_tickets.filter(priority__in=[TicketPriority.HIGH, TicketPriority.CRITICAL]).count()
        context['overdue_sla_tickets'] = active_tickets.filter(due_at__lt=timezone.now()).count()
        context['critical_open_tickets'] = active_tickets.filter(priority=TicketPriority.CRITICAL).count()
        
        stale_cutoff = timezone.now() - timedelta(days=3)
        context['stale_tickets'] = active_tickets.filter(updated_at__lte=stale_cutoff).count()
        context['assigned_tickets'] = queryset.filter(assigned_to=user).count() if user.is_technician else queryset.filter(assigned_to__isnull=False).count()
        context['recent_tickets'] = queryset.order_by('-updated_at')[:5]

        if user.is_admin:
            context['audit_count'] = AuditLog.objects.count()
            context['recent_audits'] = AuditLog.objects.select_related('user', 'ticket').order_by('-action_time')[:5]
            context['stale_notifications_available'] = context['stale_tickets'] > 0
            context['overdue_notifications_available'] = context['overdue_sla_tickets'] > 0

        # Store and Property Management context for the dashboard
        if user.is_property_manager or user.is_admin:
            context['pending_store_requests'] = MaterialRequest.objects.filter(status=MaterialRequestStatus.PENDING).count()
            context['low_stock_count'] = Material.objects.filter(stock_quantity__lte=F('min_stock_level')).count()
        
        context['my_material_requests_count'] = MaterialRequest.objects.filter(requester=user).count()

        return context


class UserListView(LoginRequiredMixin, ListView):
    model = CustomUser
    template_name = 'core/user_list.html'
    context_object_name = 'users'
    login_url = reverse_lazy('login')

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_admin:
            return HttpResponse('Access denied', status=403)
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return CustomUser.objects.order_by('username')


class UserUpdateView(LoginRequiredMixin, UpdateView):
    model = CustomUser
    form_class = UserRoleForm
    template_name = 'core/user_form.html'
    context_object_name = 'user_obj'
    success_url = reverse_lazy('user_list')
    login_url = reverse_lazy('login')

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_admin:
            return HttpResponse('Access denied', status=403)
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        response = super().form_valid(form)
        AuditLog.objects.create(
            user=self.request.user,
            action_type='user_role_updated',
            description=f"User '{self.object.username}' role updated to {self.object.role} by {self.request.user.username}.",
            ip_address=get_client_ip(self.request),
            user_agent=self.request.META.get('HTTP_USER_AGENT', ''),
            metadata={'user_id': self.object.pk, 'role': self.object.role, 'is_active': self.object.is_active},
        )
        return response


class TicketCategorySuggestionView(LoginRequiredMixin, View):
    """
    AJAX endpoint to provide AI-suggested categories as the user types.
    """
    def get(self, request, *args, **kwargs):
        title = request.GET.get('title', '')
        description = request.GET.get('description', '')
        suggestion = predict_ticket_category(title, description)
        return JsonResponse({'suggestion': suggestion})


class TicketCreateView(LoginRequiredMixin, CreateView):
    model = Ticket
    form_class = TicketForm
    template_name = 'core/ticket_form.html'
    success_url = reverse_lazy('ticket_list')
    login_url = reverse_lazy('login')

    def form_valid(self, form):
        form.instance.reported_by = self.request.user
        
        # Fallback AI-driven category suggestion if not already provided by the frontend
        if not form.instance.ai_suggested_category and form.instance.title and form.instance.description:
            ai_category = predict_ticket_category(form.instance.title, form.instance.description)
            if ai_category:
                form.instance.ai_suggested_category = ai_category

        response = super().form_valid(form)
        AuditLog.objects.create(
            user=self.request.user,
            ticket=self.object,
            action_type='ticket_created',
            description=f'Ticket "{self.object.title}" created by {self.request.user.username}.',
            ip_address=get_client_ip(self.request),
            user_agent=self.request.META.get('HTTP_USER_AGENT', ''),
            metadata={'status': self.object.status, 'priority': self.object.priority},
        )
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_create_view'] = True # Used in template for JS logic
        return context

class TicketListView(LoginRequiredMixin, ListView):
    model = Ticket
    template_name = 'core/ticket_list.html'
    context_object_name = 'tickets'
    paginate_by = 20
    login_url = reverse_lazy('login')

    def get_queryset(self):
        user = self.request.user
        queryset = Ticket.objects.select_related('reported_by', 'assigned_to')
        if user.is_admin:
            queryset = queryset.all()
        elif user.is_technician:
            queryset = queryset.filter(
                Q(assigned_to=user) | Q(status=TicketStatus.OPEN) | Q(reported_by=user)
            )
        else:
            queryset = queryset.filter(reported_by=user)

        status = self.request.GET.get('status')
        priority = self.request.GET.get('priority')
        wing = self.request.GET.get('wing')
        floor = self.request.GET.get('floor')
        assigned_to = self.request.GET.get('assigned_to')
        search = self.request.GET.get('search')

        if status:
            queryset = queryset.filter(status=status)
        if priority:
            queryset = queryset.filter(priority=priority)
        if wing:
            queryset = queryset.filter(wing=wing)
        if floor:
            queryset = queryset.filter(floor=floor)
        if assigned_to:
            queryset = queryset.filter(assigned_to__username__icontains=assigned_to)
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search) |
                Q(device_serial_number__icontains=search) |
                Q(category__icontains=search)
            )

        stale = self.request.GET.get('stale')
        if stale == '1':
            cutoff = timezone.now() - timedelta(days=3)
            queryset = queryset.filter(status__in=[TicketStatus.OPEN, TicketStatus.IN_PROGRESS], updated_at__lte=cutoff)
            
        overdue = self.request.GET.get('overdue')
        if overdue == '1':
            queryset = queryset.filter(status__in=[TicketStatus.OPEN, TicketStatus.IN_PROGRESS], due_at__lt=timezone.now())

        return queryset.order_by('-updated_at')

    def get_context_data(self, **kwargs):
        context = ListView.get_context_data(self, **kwargs)
        context['status_choices'] = TicketStatus.choices
        context['priority_choices'] = TicketPriority.choices
        context['wing_choices'] = WingChoices.choices
        context['floor_choices'] = FloorChoices.choices
        context['assigned_to_filter'] = self.request.GET.get('assigned_to', '')
        context['status_filter'] = self.request.GET.get('status', '')
        context['priority_filter'] = self.request.GET.get('priority', '')
        context['wing_filter'] = self.request.GET.get('wing', '')
        context['floor_filter'] = self.request.GET.get('floor', '')
        context['stale_filter'] = self.request.GET.get('stale', '')
        context['overdue_filter'] = self.request.GET.get('overdue', '')
        context['search_query'] = self.request.GET.get('search', '')
        return context


class TicketDetailView(LoginRequiredMixin, DetailView):
    model = Ticket
    template_name = 'core/ticket_detail.html'
    context_object_name = 'ticket'
    login_url = reverse_lazy('login')

    def get_queryset(self):
        user = self.request.user
        queryset = Ticket.objects.select_related('reported_by', 'assigned_to')
        if user.is_admin:
            return queryset.all()
        if user.is_technician:
            return queryset.filter(
                Q(assigned_to=user) | Q(status=TicketStatus.OPEN) | Q(reported_by=user)
            )
        return queryset.filter(reported_by=user)

    def get_context_data(self, **kwargs):
        context = DetailView.get_context_data(self, **kwargs)
        context['audit_logs'] = self.object.audit_logs.select_related('user').all()[:10]
        context['comments'] = self.object.comments.select_related('user').order_by('created_at')
        context['attachments'] = self.object.attachments.select_related('uploaded_by').order_by('-uploaded_at')
        context['comment_form'] = TicketCommentForm()
        context['attachment_form'] = TicketAttachmentForm()
        context['reopen_form'] = TicketReopenForm()
        return context


class TicketAssignView(LoginRequiredMixin, View):
    login_url = reverse_lazy('login')

    def post(self, request, pk, *args, **kwargs):
        ticket = get_object_or_404(Ticket.objects.select_related('reported_by', 'assigned_to'), pk=pk)
        if not (request.user.is_admin or request.user.is_technician):
            return HttpResponse('Access denied', status=403)
        if ticket.assigned_to and ticket.assigned_to != request.user:
            return HttpResponse('Ticket already assigned', status=403)

        ticket.assigned_to = request.user
        if ticket.status == TicketStatus.OPEN:
            ticket.status = TicketStatus.IN_PROGRESS
        ticket.save()

        AuditLog.objects.create(
            user=request.user,
            ticket=ticket,
            action_type='ticket_assigned',
            description=f'Ticket "{ticket.title}" assigned to {request.user.username}.',
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            metadata={'assigned_to': request.user.username},
        )

        if ticket.assigned_to and ticket.assigned_to.email:
            send_notification_email(
                subject=f'EAC Helpdesk: Ticket assigned to you',
                message=f"Ticket '{ticket.title}' has been assigned to you. View it here: {request.build_absolute_uri(reverse_lazy('ticket_detail', kwargs={'pk': ticket.pk}))}",
                recipients=[ticket.assigned_to.email],
            )
        if ticket.reported_by and ticket.reported_by.email:
            send_notification_email(
                subject=f'EAC Helpdesk: Ticket assigned',
                message=f"Your ticket '{ticket.title}' has been assigned to {ticket.assigned_to.get_full_name() or ticket.assigned_to.username}.",
                recipients=[ticket.reported_by.email],
            )
        return redirect('ticket_detail', pk=pk)


class TicketResolveView(LoginRequiredMixin, View):
    login_url = reverse_lazy('login')

    def post(self, request, pk, *args, **kwargs):
        ticket = get_object_or_404(Ticket.objects.select_related('reported_by', 'assigned_to'), pk=pk)
        if not (request.user.is_admin or request.user.is_technician):
            return HttpResponse('Access denied', status=403)
        if request.user.is_technician and ticket.assigned_to != request.user:
            return HttpResponse('Only the assigned technician can resolve this ticket.', status=403)
        if ticket.status == TicketStatus.RESOLVED:
            return redirect('ticket_detail', pk=pk)

        ticket.status = TicketStatus.RESOLVED
        ticket.save()

        AuditLog.objects.create(
            user=request.user,
            ticket=ticket,
            action_type='ticket_resolved',
            description=f'Ticket "{ticket.title}" marked as resolved by {request.user.username}.',
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            metadata={'resolved_by': request.user.username},
        )

        if ticket.reported_by and ticket.reported_by.email:
            send_notification_email(
                subject=f'EAC Helpdesk: Ticket resolved',
                message=f"Your ticket '{ticket.title}' has been marked as resolved by {request.user.get_full_name() or request.user.username}.",
                recipients=[ticket.reported_by.email],
            )
        if ticket.assigned_to and ticket.assigned_to.email:
            send_notification_email(
                subject=f'EAC Helpdesk: Ticket you resolved',
                message=f"You marked ticket '{ticket.title}' as resolved.",
                recipients=[ticket.assigned_to.email],
            )
        return redirect('ticket_detail', pk=pk)


class TicketCloseView(LoginRequiredMixin, View):
    login_url = reverse_lazy('login')

    def post(self, request, pk, *args, **kwargs):
        ticket = get_object_or_404(Ticket.objects.select_related('reported_by', 'assigned_to'), pk=pk)
        if not (request.user.is_admin or request.user.is_technician):
            return HttpResponse('Access denied', status=403)
        if request.user.is_technician and ticket.assigned_to != request.user:
            return HttpResponse('Only the assigned technician can close this ticket.', status=403)
        if ticket.status == TicketStatus.CLOSED:
            return redirect('ticket_detail', pk=pk)

        ticket.status = TicketStatus.CLOSED
        ticket.save()

        AuditLog.objects.create(
            user=request.user,
            ticket=ticket,
            action_type='ticket_closed',
            description=f'Ticket "{ticket.title}" closed by {request.user.username}.',
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            metadata={'closed_by': request.user.username},
        )

        if ticket.reported_by and ticket.reported_by.email:
            send_notification_email(
                subject=f'EAC Helpdesk: Ticket closed',
                message=f"Your ticket '{ticket.title}' has been closed by {request.user.get_full_name() or request.user.username}.",
                recipients=[ticket.reported_by.email],
            )
        return redirect('ticket_detail', pk=pk)


class TicketReopenView(LoginRequiredMixin, View):
    login_url = reverse_lazy('login')

    def post(self, request, pk, *args, **kwargs):
        ticket = get_object_or_404(Ticket.objects.select_related('reported_by', 'assigned_to'), pk=pk)
        if not (request.user.is_admin or request.user.is_technician):
            return HttpResponse('Access denied', status=403)
        if ticket.status not in (TicketStatus.RESOLVED, TicketStatus.CLOSED):
            return redirect('ticket_detail', pk=pk)

        form = TicketReopenForm(request.POST)
        if not form.is_valid():
            context = {
                'ticket': ticket,
                'audit_logs': ticket.audit_logs.select_related('user').all()[:10],
                'comments': ticket.comments.select_related('user').order_by('created_at'),
                'attachments': ticket.attachments.select_related('uploaded_by').order_by('-uploaded_at'),
                'comment_form': TicketCommentForm(),
                'attachment_form': TicketAttachmentForm(),
                'reopen_form': form,
            }
            return render(request, 'core/ticket_detail.html', context)

        previous_status = ticket.status
        ticket.status = TicketStatus.IN_PROGRESS if ticket.assigned_to else TicketStatus.OPEN
        ticket.resolved_at = None
        ticket.closed_at = None
        ticket.reopened_at = timezone.now()
        ticket.reopen_reason = form.cleaned_data.get('reason', '')
        ticket.save()

        AuditLog.objects.create(
            user=request.user,
            ticket=ticket,
            action_type='ticket_reopened',
            description=f'Ticket "{ticket.title}" reopened by {request.user.username} from {previous_status}.',
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            metadata={
                'reopened_by': request.user.username,
                'previous_status': previous_status,
                'reopen_reason': ticket.reopen_reason,
            },
        )

        if ticket.reported_by and ticket.reported_by.email:
            send_notification_email(
                subject=f'EAC Helpdesk: Ticket reopened',
                message=f"Your ticket '{ticket.title}' has been reopened by {request.user.get_full_name() or request.user.username}.",
                recipients=[ticket.reported_by.email],
            )
        return redirect('ticket_detail', pk=pk)


class TicketCommentCreateView(LoginRequiredMixin, View):
    login_url = reverse_lazy('login')

    def post(self, request, pk, *args, **kwargs):
        ticket = get_object_or_404(Ticket.objects.select_related('reported_by', 'assigned_to'), pk=pk)
        if not user_can_access_ticket(request.user, ticket):
            return HttpResponse('Access denied', status=403)
        if not user_can_modify_ticket(request.user, ticket):
            return HttpResponse('Access denied', status=403)

        form = TicketCommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.ticket = ticket
            comment.user = request.user
            comment.save()
            AuditLog.objects.create(
                user=request.user,
                ticket=ticket,
                action_type='comment_added',
                description=f'Comment added to ticket "{ticket.title}" by {request.user.username}.',
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                metadata={'comment_id': comment.pk},
            )
            return redirect('ticket_detail', pk=pk)

        context = {
            'ticket': ticket,
            'audit_logs': ticket.audit_logs.select_related('user').all()[:10],
            'comments': ticket.comments.select_related('user').order_by('created_at'),
            'attachments': ticket.attachments.select_related('uploaded_by').order_by('-uploaded_at'),
            'comment_form': form,
            'attachment_form': TicketAttachmentForm(),
        }
        return render(request, 'core/ticket_detail.html', context)


class TicketAttachmentCreateView(LoginRequiredMixin, View):
    login_url = reverse_lazy('login')

    def post(self, request, pk, *args, **kwargs):
        ticket = get_object_or_404(Ticket.objects.select_related('reported_by', 'assigned_to'), pk=pk)
        if not user_can_access_ticket(request.user, ticket):
            return HttpResponse('Access denied', status=403)
        if not user_can_modify_ticket(request.user, ticket):
            return HttpResponse('Access denied', status=403)

        form = TicketAttachmentForm(request.POST, request.FILES)
        if form.is_valid():
            attachment = form.save(commit=False)
            attachment.ticket = ticket
            attachment.uploaded_by = request.user
            attachment.save()
            AuditLog.objects.create(
                user=request.user,
                ticket=ticket,
                action_type='attachment_uploaded',
                description=f'Attachment uploaded to ticket "{ticket.title}" by {request.user.username}.',
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                metadata={'attachment_id': attachment.pk, 'filename': attachment.file.name},
            )
            return redirect('ticket_detail', pk=pk)

        context = {
            'ticket': ticket,
            'audit_logs': ticket.audit_logs.select_related('user').all()[:10],
            'comments': ticket.comments.select_related('user').order_by('created_at'),
            'comment_form': TicketCommentForm(),
            'attachments': ticket.attachments.select_related('uploaded_by').order_by('-uploaded_at'),
            'attachment_form': form,
        }
        return render(request, 'core/ticket_detail.html', context)


class TicketUpdateView(LoginRequiredMixin, UpdateView):
    model = Ticket
    form_class = TicketUpdateForm
    template_name = 'core/ticket_update.html'
    context_object_name = 'ticket'
    login_url = reverse_lazy('login')

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if not user_can_modify_ticket(request.user, self.object):
            return HttpResponse('Access denied', status=403)
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        user = self.request.user
        queryset = Ticket.objects.select_related('reported_by', 'assigned_to')
        if user.is_admin:
            return queryset.all()
        if user.is_technician:
            return queryset.filter(
                Q(assigned_to=user) | Q(reported_by=user)
            )
        return queryset.filter(reported_by=user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        changed_fields = form.changed_data
        response = super().form_valid(form)
        AuditLog.objects.create(
            user=self.request.user,
            ticket=self.object,
            action_type='ticket_updated',
            description=f'Ticket "{self.object.title}" updated by {self.request.user.username}. Changed fields: {", ".join(changed_fields)}',
            ip_address=get_client_ip(self.request),
            user_agent=self.request.META.get('HTTP_USER_AGENT', ''),
            metadata={'changed_fields': changed_fields},
        )
        return response

    def get_success_url(self):
        return reverse_lazy('ticket_detail', kwargs={'pk': self.object.pk})


class TicketReportView(LoginRequiredMixin, TemplateView):
    template_name = 'core/ticket_report.html'
    login_url = reverse_lazy('login')

    def dispatch(self, request, *args, **kwargs):
        if not (request.user.is_admin or request.user.is_technician):
            return HttpResponse('Access denied', status=403)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = kwargs.copy()
        context['view'] = self
        queryset = apply_ticket_report_filters(
            get_report_ticket_queryset(self.request.user),
            self.request.GET,
        )

        context['status_filter'] = self.request.GET.get('status', '')
        context['priority_filter'] = self.request.GET.get('priority', '')
        context['category_filter'] = self.request.GET.get('category', '')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        context['status_choices'] = TicketStatus.choices
        context['priority_choices'] = TicketPriority.choices

        context['open_count'] = queryset.filter(status=TicketStatus.OPEN).count()
        context['in_progress_count'] = queryset.filter(status=TicketStatus.IN_PROGRESS).count()
        stale_cutoff = timezone.now() - timedelta(days=3)
        context['stale_count'] = queryset.filter(status__in=[TicketStatus.OPEN, TicketStatus.IN_PROGRESS], updated_at__lte=stale_cutoff).count()
        context['overdue_count'] = queryset.filter(status__in=[TicketStatus.OPEN, TicketStatus.IN_PROGRESS], due_at__lt=timezone.now()).count()
        context['resolved_count'] = queryset.filter(status=TicketStatus.RESOLVED).count()
        context['closed_count'] = queryset.filter(status=TicketStatus.CLOSED).count()
        context['priority_counts'] = queryset.values('priority').annotate(count=Count('priority')).order_by('-count')
        context['category_counts'] = queryset.values('category').annotate(count=Count('id')).order_by('-count')
        context['recent_tickets'] = queryset.order_by('-updated_at')[:6]
        return context


class SendStaleTicketNotificationsView(LoginRequiredMixin, UserPassesTestMixin, View):
    login_url = reverse_lazy('login')

    def test_func(self):
        return self.request.user.is_admin

    def handle_no_permission(self):
        return HttpResponse('Access denied', status=403)

    def post(self, request, *args, **kwargs):
        sent_count, recipient_count = send_stale_ticket_notification_emails()
        if sent_count:
            messages.success(request, f'Stale ticket alerts sent for {sent_count} ticket(s) to {recipient_count} recipient(s).')
        else:
            messages.info(request, 'No stale tickets were found, so no alerts were sent.')
        return redirect('dashboard')


class SendOverdueTicketNotificationsView(LoginRequiredMixin, UserPassesTestMixin, View):
    login_url = reverse_lazy('login')

    def test_func(self):
        return self.request.user.is_admin

    def handle_no_permission(self):
        return HttpResponse('Access denied', status=403)

    def post(self, request, *args, **kwargs):
        sent_count, recipient_count = send_overdue_ticket_notification_emails()
        if sent_count:
            messages.success(request, f'Overdue ticket alerts sent for {sent_count} ticket(s) to {recipient_count} recipient(s).')
        else:
            messages.info(request, 'No overdue tickets were found, so no alerts were sent.')
        return redirect('dashboard')


class AuditLogPDFView(LoginRequiredMixin, UserPassesTestMixin, View):
    login_url = reverse_lazy('login')

    def test_func(self):
        return self.request.user.is_admin

    def handle_no_permission(self):
        return HttpResponse('Access denied', status=403)

    def get(self, request, *args, **kwargs):
        audit_queryset = AuditLog.objects.select_related('user', 'ticket').order_by('-action_time')
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, title='Audit Report')
        styles = getSampleStyleSheet()
        elements = [
            Paragraph('EAC Helpdesk Audit Report', styles['Title']),
            Spacer(1, 12),
            Paragraph(f'Generated by: {request.user.get_full_name() or request.user.username}', styles['Normal']),
            Paragraph(f'Date: {timezone.now():%Y-%m-%d %H:%M:%S}', styles['Normal']),
            Spacer(1, 12),
            Paragraph('Audit events', styles['Heading2']),
        ]

        table_data = [[
            'Time', 'User', 'Action', 'Ticket', 'IP Address', 'Description'
        ]]
        for event in audit_queryset[:100]:
            table_data.append([
                event.action_time.strftime('%Y-%m-%d %H:%M:%S'),
                event.user.get_full_name() if event.user else 'System',
                event.action_type,
                event.ticket.title if event.ticket else 'System',
                event.ip_address or '-',
                event.description,
            ])

        table = Table(table_data, repeatRows=1, hAlign='LEFT')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ]))
        elements.append(table)

        doc.build(elements)
        AuditLog.objects.create(
            user=request.user,
            action_type='audit_exported',
            description=f'Audit log exported as PDF by {request.user.username}.',
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            metadata={'format': 'pdf', 'event_count': audit_queryset.count()},
        )
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="audit_report.pdf"'
        return response


class AuditLogExcelView(LoginRequiredMixin, UserPassesTestMixin, View):
    login_url = reverse_lazy('login')

    def test_func(self):
        return self.request.user.is_admin

    def handle_no_permission(self):
        return HttpResponse('Access denied', status=403)

    def get(self, request, *args, **kwargs):
        audit_queryset = AuditLog.objects.select_related('user', 'ticket').order_by('-action_time')
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Audit Report'

        sheet.append(['EAC Helpdesk Audit Report'])
        sheet.append([f'Generated by: {request.user.get_full_name() or request.user.username}'])
        sheet.append([f'Date: {timezone.now():%Y-%m-%d %H:%M:%S}'])
        sheet.append([])
        sheet.append(['Time', 'User', 'Action', 'Ticket', 'IP Address', 'Description'])

        for event in audit_queryset:
            sheet.append([
                event.action_time.strftime('%Y-%m-%d %H:%M:%S'),
                event.user.get_full_name() if event.user else 'System',
                event.action_type,
                event.ticket.title if event.ticket else 'System',
                event.ip_address or '-',
                event.description,
            ])

        for idx, column_width in enumerate([20, 24, 20, 30, 18, 60], start=1):
            sheet.column_dimensions[sheet.cell(row=5, column=idx).column_letter].width = column_width

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        AuditLog.objects.create(
            user=request.user,
            action_type='audit_exported',
            description=f'Audit log exported as Excel by {request.user.username}.',
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            metadata={'format': 'excel', 'event_count': audit_queryset.count()},
        )
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="audit_report.xlsx"'
        return response


class AuditLogCSVView(LoginRequiredMixin, UserPassesTestMixin, View):
    login_url = reverse_lazy('login')

    def test_func(self):
        return self.request.user.is_admin

    def handle_no_permission(self):
        return HttpResponse('Access denied', status=403)

    def get(self, request, *args, **kwargs):
        audit_queryset = AuditLog.objects.select_related('user', 'ticket').order_by('-action_time')
        csv_buffer = StringIO()
        writer = csv.writer(csv_buffer)
        writer.writerow(['Time', 'User', 'Action', 'Ticket', 'IP Address', 'Description'])
        for event in audit_queryset:
            writer.writerow([
                event.action_time.strftime('%Y-%m-%d %H:%M:%S'),
                event.user.get_full_name() if event.user else 'System',
                event.action_type,
                event.ticket.title if event.ticket else 'System',
                event.ip_address or '-',
                event.description,
            ])
        AuditLog.objects.create(
            user=request.user,
            action_type='audit_exported',
            description=f'Audit log exported as CSV by {request.user.username}.',
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            metadata={'format': 'csv', 'event_count': audit_queryset.count()},
        )
        response = HttpResponse(csv_buffer.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="audit_report.csv"'
        return response


class AuditLogListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = AuditLog
    template_name = 'core/audit_log_list.html'
    context_object_name = 'audit_logs'
    paginate_by = 25
    login_url = reverse_lazy('login')

    def test_func(self):
        return self.request.user.is_admin

    def handle_no_permission(self):
        return HttpResponse('Access denied', status=403)

    def get_queryset(self):
        queryset = AuditLog.objects.select_related('user', 'ticket').all()
        return apply_audit_filters(queryset, self.request.GET)

    def get_context_data(self, **kwargs):
        context = ListView.get_context_data(self, **kwargs)
        queryset = self.get_queryset()
        context['total_audit_events'] = queryset.count()
        context['top_action_types'] = queryset.values('action_type').annotate(count=Count('action_type')).order_by('-count')[:5]
        context['action_type_filter'] = self.request.GET.get('action_type', '')
        context['user_filter'] = self.request.GET.get('user', '')
        context['ticket_filter'] = self.request.GET.get('ticket', '')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        context['available_action_types'] = AuditLog.objects.values_list('action_type', flat=True).distinct().order_by('action_type')
        return context


class TicketReportPDFView(LoginRequiredMixin, View):
    login_url = reverse_lazy('login')

    def dispatch(self, request, *args, **kwargs):
        if not (request.user.is_admin or request.user.is_technician):
            return HttpResponse('Access denied', status=403)
        return super().dispatch(request, *args, **kwargs)

    def _auto_recommendations(self, stats: dict) -> list[str]:
        recs = []
        total = stats.get('total_incidents', 0) or 0
        avg_hours = stats.get('avg_resolution_hours')
        most_cat = stats.get('most_common_category')
        most_pri = stats.get('most_common_priority')
        top_dept = stats.get('top_department')

        if total:
            recs.append('Maintain continuous monitoring of open and in-progress incidents to prevent aging and escalation.')

        if avg_hours is not None and not math.isnan(avg_hours):
            if avg_hours > 48:
                recs.append('Average resolution time is above 48 hours. Prioritize workflow improvements and escalation for overdue cases.')
            else:
                recs.append('Average resolution time is within an acceptable range. Continue current triage and resolution practices.')
        else:
            recs.append('Resolution-time data is incomplete for some incidents. Ensure resolved timestamps are consistently recorded.')

        if most_cat:
            recs.append(f"Category '{most_cat}' is most common. Implement targeted preventive actions (training, spares, or process improvements) for this category.")
        if most_pri == TicketPriority.CRITICAL:
            recs.append('Critical incidents detected. Conduct rapid RCA (root-cause analysis) and ensure immediate containment actions are followed.')

        if top_dept:
            recs.append(f"Department/Wing with highest incident volume: '{top_dept}'. Review departmental controls and address recurring risk areas.")

        recs.append('Ensure audit trail completeness by recording all status transitions and resolutions with the acting user.')
        return recs

    def get(self, request, *args, **kwargs):
        queryset = apply_ticket_report_filters(get_report_ticket_queryset(request.user), request.GET)

        # Preload data to reduce query count.
        queryset = queryset.select_related('reported_by', 'assigned_to').all()
        tickets = list(queryset.order_by('-updated_at'))

        total_incidents = len(tickets)
        open_incidents = sum(1 for t in tickets if t.status == TicketStatus.OPEN)
        in_progress_incidents = sum(1 for t in tickets if t.status == TicketStatus.IN_PROGRESS)
        resolved_incidents = sum(1 for t in tickets if t.status == TicketStatus.RESOLVED)
        closed_incidents = sum(1 for t in tickets if t.status == TicketStatus.CLOSED)

        resolution_durations = []
        for t in tickets:
            if t.resolved_at:
                resolution_durations.append((t.resolved_at - t.created_at).total_seconds() / 3600)
        avg_resolution_hours = (sum(resolution_durations) / len(resolution_durations)) if resolution_durations else float('nan')

        # Statistics breakdowns
        incidents_by_department = {}
        incidents_by_priority = {}
        incidents_by_status = {}
        category_counts = {}

        for t in tickets:
            incidents_by_department[t.wing or 'Not specified'] = incidents_by_department.get(t.wing or 'Not specified', 0) + 1
            incidents_by_priority[t.priority] = incidents_by_priority.get(t.priority, 0) + 1
            incidents_by_status[t.status] = incidents_by_status.get(t.status, 0) + 1
            key_cat = t.category or 'Uncategorized'
            category_counts[key_cat] = category_counts.get(key_cat, 0) + 1

        most_common_category = max(category_counts.items(), key=lambda kv: kv[1])[0] if category_counts else None
        most_common_priority = max(incidents_by_priority.items(), key=lambda kv: kv[1])[0] if incidents_by_priority else None
        top_department = max(incidents_by_department.items(), key=lambda kv: kv[1])[0] if incidents_by_department else None

        stats = {
            'total_incidents': total_incidents,
            'open_incidents': open_incidents,
            'in_progress_incidents': in_progress_incidents,
            'resolved_incidents': resolved_incidents,
            'closed_incidents': closed_incidents,
            'avg_resolution_hours': avg_resolution_hours,
            'incidents_by_department': incidents_by_department,
            'incidents_by_priority': incidents_by_priority,
            'incidents_by_status': incidents_by_status,
            'most_common_category': most_common_category,
            'most_common_priority': most_common_priority,
            'top_department': top_department,
        }

        # Charts (PDF embeds image bytes)
        chart_images = []  # list[Image]
        if plt is not None and total_incidents > 0:
            def _render_pie(title: str, data: dict):
                labels = list(data.keys())
                values = list(data.values())
                fig = plt.figure(figsize=(6, 3.5))
                ax = fig.add_subplot(111)
                ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90)
                ax.set_title(title)
                buf = BytesIO()
                fig.tight_layout()
                fig.savefig(buf, format='png', dpi=160)
                plt.close(fig)
                buf.seek(0)
                return Image(buf, width=5.7*inch, height=3.3*inch)

            chart_images.append(_render_pie('Incidents by Department (Wing)', incidents_by_department))
            chart_images.append(_render_pie('Incidents by Priority', {k: v for k, v in incidents_by_priority.items()}))
            chart_images.append(_render_pie('Incidents by Status', {k: v for k, v in incidents_by_status.items()}))

        # Resolution details + audit trail (best-effort)
        ticket_ids = [t.id for t in tickets]
        audit_logs = (AuditLog.objects.select_related('user', 'ticket')
                       .filter(ticket_id__in=ticket_ids)
                       .order_by('ticket_id', 'action_time'))
        audit_by_ticket = {}
        for a in audit_logs:
            audit_by_ticket.setdefault(a.ticket_id, []).append(a)

        # Pre-fetch only the devices relevant to the tickets in this report (using PK for uniqueness)
        relevant_device_ids = [t.device_id for t in tickets if t.device_id]
        devices_map = {d.pk: d for d in Device.objects.filter(pk__in=relevant_device_ids)}
        
        comments = (TicketComment.objects.filter(ticket_id__in=ticket_ids)
                    .select_related('user')
                    .order_by('ticket_id', '-created_at'))
        latest_comment = {}
        for c in comments:
            if c.ticket_id not in latest_comment:
                latest_comment[c.ticket_id] = c

        # PDF document with footer page numbers
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, title='EAC Incident Audit Report')
        styles = getSampleStyleSheet()
        elements = []

        def _draw_page_number(canvas: Canvas, doc_obj):
            canvas.saveState()
            canvas.setFont('Helvetica', 9)
            canvas.drawRightString(letter[0] - inch, 0.6*inch/2, f"Page {doc_obj.page}")
            canvas.restoreState()

        # Header: logo + branding
        logo_path = getattr(settings, 'MEDIA_ROOT', None)
        logo_img = None
        if logo_path:
            candidate = [
                os.path.join(logo_path, 'east-african-community-logo-png_seeklogo-511714.png'),
                os.path.join(logo_path, 'EAC_Logo.png'),
            ]
            for p in candidate:
                if os.path.exists(p):
                    logo_img = p
                    break

        if logo_img:
            try:
                img = Image(logo_img, width=1.0*inch, height=1.0*inch, hAlign='CENTER')
                elements.append(img)
            except Exception:
                pass

        styles['h2'].alignment = 1  # Set hAlign to center for the organization name
        elements.append(Paragraph('East African Community', styles['h2'])) # Added line
        elements += [
            Paragraph('EAC Helpdesk Incident / Audit Report', styles['Title']),
            Paragraph(f"Report generated: {timezone.now():%Y-%m-%d %H:%M:%S}", styles['Normal']),
            Paragraph(f"Prepared/Generated By: {request.user.get_full_name() or request.user.username}", styles['Normal']),
            Spacer(1, 10),
            Paragraph('Executive Summary', styles['Heading2']),
        ]

        summary_data = [
            ['Total Incidents', str(total_incidents)],
            ['Open Incidents', str(open_incidents)],
            ['In Progress Incidents', str(in_progress_incidents)],
            ['Resolved Incidents', str(resolved_incidents)],
            ['Closed Incidents', str(closed_incidents)],
            ['Average Resolution Time (hours)', '—' if math.isnan(avg_resolution_hours) else f"{avg_resolution_hours:.2f}"],
        ]
        summary_table = Table(summary_data, colWidths=[250, 90])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(summary_table)

        elements.append(Spacer(1, 12))
        elements.append(Paragraph('Statistics & Charts', styles['Heading2']))
        for img in chart_images:
            elements.append(img)
            elements.append(Spacer(1, 8))

        elements.append(Paragraph('Detailed Incident Table', styles['Heading2']))

        table_header = [
            'Ticket #', 'Incident Title', 'Description', 'Department (Wing)', 'Office Location',
            'Reported By', 'Assigned Technician', 'Priority', 'Status',
            'Date Reported', 'Date Resolved',
        ]
        detailed_rows = [table_header]

        for t in tickets:
            office_location = ', '.join([x for x in [t.floor or None, t.room_number or None] if x]) or '—'
            detailed_rows.append([
                str(t.id),
                t.title,
                t.description or '—',
                t.wing or 'Not specified',
                office_location,
                (t.reported_by.get_full_name() or t.reported_by.username) if t.reported_by_id else '—',
                (t.assigned_to.get_full_name() if t.assigned_to else 'Unassigned'),
                t.get_priority_display(),
                t.get_status_display(),
                t.created_at.strftime('%Y-%m-%d'),
                t.resolved_at.strftime('%Y-%m-%d') if t.resolved_at else '—',
            ])

        detailed_table = Table(detailed_rows, repeatRows=1, hAlign='LEFT')
        detailed_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
        ]))
        elements.append(detailed_table)

        elements.append(Spacer(1, 12))
        elements.append(Paragraph('Resolution Details', styles['Heading2']))

        res_rows = [['Ticket #', 'Resolution Notes', 'Resolved By', 'Resolution Date']]
        for t in tickets:
            notes = '—'
            if t.id in latest_comment:
                c = latest_comment[t.id]
                notes = c.comment[:500] if c.comment else '—'

            resolved_by = '—'
            res_date = t.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if t.resolved_at else '—'

            logs = audit_by_ticket.get(t.id, [])
            resolved_log = next((a for a in logs if a.action_type == 'ticket_resolved'), None)
            if resolved_log and resolved_log.user:
                resolved_by = resolved_log.user.get_full_name() or resolved_log.user.username
            elif t.status in (TicketStatus.RESOLVED, TicketStatus.CLOSED):
                # best-effort fallback: the latest audit log user for the ticket
                if logs and logs[-1].user:
                    resolved_by = logs[-1].user.get_full_name() or logs[-1].user.username

            res_rows.append([str(t.id), notes, resolved_by, res_date])

        res_table = Table(res_rows, repeatRows=1, hAlign='LEFT')
        res_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(res_table)

        elements.append(Spacer(1, 12))
        elements.append(Paragraph('Audit Trail', styles['Heading2']))

        audit_rows = [['Ticket #', 'User', 'Action Performed', 'Date and Time', 'Previous Status', 'New Status']]
        for t in tickets:
            logs = audit_by_ticket.get(t.id, [])
            for a in logs[:200]:
                prev_status = (a.metadata or {}).get('previous_status')
                new_status = None
                if a.action_type == 'ticket_assigned':
                    new_status = 'In Progress' if t.status == TicketStatus.IN_PROGRESS else None
                if a.action_type == 'ticket_resolved':
                    new_status = 'Resolved'
                if a.action_type == 'ticket_closed':
                    new_status = 'Closed'

                audit_rows.append([
                    str(t.id),
                    (a.user.get_full_name() if a.user else 'System'),
                    a.action_type,
                    a.action_time.strftime('%Y-%m-%d %H:%M:%S'),
                    prev_status or '—',
                    new_status or '—',
                ])

        audit_table = Table(audit_rows, repeatRows=1, hAlign='LEFT')
        audit_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(audit_table)

        elements.append(Spacer(1, 12))
        elements.append(Paragraph('Recommendations', styles['Heading2']))
        recs = self._auto_recommendations(stats)
        for r in recs:
            elements.append(Paragraph(f"• {r}", styles['Normal']))

        elements.append(Spacer(1, 18))
        elements.append(Paragraph('Approval', styles['Heading2']))
        approval_table = Table([
            ['Prepared By', 'Reviewed By', 'Approved By', 'Signature', 'Date'],
            [request.user.get_full_name() or request.user.username, '—', '—', '_________________', timezone.now().strftime('%Y-%m-%d')],
        ], colWidths=[120, 120, 120, 120, 70])
        approval_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(approval_table)

        doc.build(elements, onFirstPage=_draw_page_number, onLaterPages=_draw_page_number)

        AuditLog.objects.create(
            user=request.user,
            action_type='ticket_exported',
            description=f'Ticket incident audit report exported as PDF by {request.user.username}.',
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            metadata={'format': 'pdf', 'ticket_count': queryset.count()},
        )

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="incident_audit_report.pdf"'
        return response



class TicketReportExcelView(LoginRequiredMixin, View):
    login_url = reverse_lazy('login')

    def dispatch(self, request, *args, **kwargs):
        if not (request.user.is_admin or request.user.is_technician):
            return HttpResponse('Access denied', status=403)
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        queryset = apply_ticket_report_filters(get_report_ticket_queryset(request.user), request.GET)
        queryset = queryset.select_related('reported_by', 'assigned_to').all()
        tickets = list(queryset.order_by('-updated_at'))
        # Pre-fetch only the devices relevant to the tickets in this report (using PK for uniqueness)
        relevant_device_ids = [t.device_id for t in tickets if t.device_id]
        devices_map = {d.pk: d for d in Device.objects.filter(pk__in=relevant_device_ids)}

        def _autosize_columns(ws, min_width=8, max_width=55):
            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    try:
                        v = '' if cell.value is None else str(cell.value)
                        max_len = max(max_len, len(v))
                    except Exception:
                        pass
                width = min(max_len + 2, max_width)
                ws.column_dimensions[col_letter].width = max(min_width, width)

        from openpyxl.styles import Font, PatternFill, Alignment

        workbook = Workbook()

        # Sheet 1: Executive Summary + Statistics
        sheet_summary = workbook.active
        sheet_summary.title = 'Executive Summary'

        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        sheet_summary.append(['EAC Helpdesk Incident / Audit Report'])
        sheet_summary.append([f"Generated by: {request.user.get_full_name() or request.user.username}"])
        sheet_summary.append([f"Generated at: {timezone.now():%Y-%m-%d %H:%M:%S}"])
        sheet_summary.append([])

        total_incidents = len(tickets)
        open_incidents = sum(1 for t in tickets if t.status == TicketStatus.OPEN)
        in_progress_incidents = sum(1 for t in tickets if t.status == TicketStatus.IN_PROGRESS)
        resolved_incidents = sum(1 for t in tickets if t.status == TicketStatus.RESOLVED)
        closed_incidents = sum(1 for t in tickets if t.status == TicketStatus.CLOSED)

        resolution_durations = []
        for t in tickets:
            if t.resolved_at:
                resolution_durations.append((t.resolved_at - t.created_at).total_seconds() / 3600)
        avg_resolution_hours = (sum(resolution_durations) / len(resolution_durations)) if resolution_durations else None

        sheet_summary.append(['Executive Summary', 'Value'])
        for cell in sheet_summary[5]:
            cell.fill = header_fill
            cell.font = header_font

        summary_map = [
            ('Total Incidents', total_incidents),
            ('Open Incidents', open_incidents),
            ('In Progress Incidents', in_progress_incidents),
            ('Resolved Incidents', resolved_incidents),
            ('Closed Incidents', closed_incidents),
            ('Average Resolution Time (hours)', '' if avg_resolution_hours is None else round(avg_resolution_hours, 2)),
        ]
        for k, v in summary_map:
            sheet_summary.append([k, v])

        # Statistics tables
        incidents_by_department = {}
        incidents_by_priority = {}
        incidents_by_status = {}
        category_counts = {}

        for t in tickets:
            incidents_by_department[t.wing or 'Not specified'] = incidents_by_department.get(t.wing or 'Not specified', 0) + 1
            incidents_by_priority[t.priority] = incidents_by_priority.get(t.priority, 0) + 1
            incidents_by_status[t.status] = incidents_by_status.get(t.status, 0) + 1
            key_cat = t.category or 'Uncategorized'
            category_counts[key_cat] = category_counts.get(key_cat, 0) + 1

        most_common_category = max(category_counts.items(), key=lambda kv: kv[1])[0] if category_counts else ''

        sheet_summary.append([])
        sheet_summary.append(['Incidents by Department (Wing)'])
        sheet_summary.append(['Department', 'Count'])
        for cell in sheet_summary[sheet_summary.max_row]:
            cell.fill = header_fill
            cell.font = header_font
        for k, v in sorted(incidents_by_department.items(), key=lambda kv: kv[1], reverse=True):
            sheet_summary.append([k, v])

        sheet_summary.append([])
        sheet_summary.append(['Incidents by Priority'])
        sheet_summary.append(['Priority', 'Count'])
        for cell in sheet_summary[sheet_summary.max_row]:
            cell.fill = header_fill
            cell.font = header_font
        for k, v in sorted(incidents_by_priority.items(), key=lambda kv: kv[1], reverse=True):
            # map to display label
            try:
                display = dict(TicketPriority.choices).get(k, k)
            except Exception:
                display = k
            sheet_summary.append([display, v])

        sheet_summary.append([])
        sheet_summary.append(['Incidents by Status'])
        sheet_summary.append(['Status', 'Count'])
        for cell in sheet_summary[sheet_summary.max_row]:
            cell.fill = header_fill
            cell.font = header_font
        for k, v in sorted(incidents_by_status.items(), key=lambda kv: kv[1], reverse=True):
            try:
                display = dict(TicketStatus.choices).get(k, k)
            except Exception:
                display = k
            sheet_summary.append([display, v])

        sheet_summary.append([])
        sheet_summary.append(['Most Common Incident Category', most_common_category])

        _autosize_columns(sheet_summary)

        # Sheet 2: Detailed Incident Table
        ws_details = workbook.create_sheet('Detailed Incidents')
        ws_details.append([
            'Ticket #', 'Incident Title', 'Device Name', 'Device Condition', 
            'Description', 'Department (Wing)', 'Office Location',
            'Reported By', 'Assigned Technician', 'Priority', 'Status',
            'Date Reported', 'Date Resolved'
        ])
        ws_details.freeze_panes = 'A2'
        for c in ws_details[1]:
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(wrap_text=True, vertical='top')

        for t in tickets:
            office_location = ', '.join([x for x in [t.floor or None, t.room_number or None] if x]) or '—'
            device_info = devices_map.get(t.device_id)
            ws_details.append([
                t.id,
                t.title,
                (device_info.name if device_info else 'N/A'),
                (device_info.get_condition_display() if device_info else 'N/A'),
                t.description or '—',
                t.wing or 'Not specified',
                office_location,
                (t.reported_by.get_full_name() or t.reported_by.username) if t.reported_by_id else '—',
                (t.assigned_to.get_full_name() if t.assigned_to else 'Unassigned'),
                t.get_priority_display(),
                t.get_status_display(),
                t.created_at.strftime('%Y-%m-%d'),
                t.resolved_at.strftime('%Y-%m-%d') if t.resolved_at else '—',
            ])

        _autosize_columns(ws_details)

        # Sheet 3: Resolution Details + Audit Trail (best-effort)
        ws_audit = workbook.create_sheet('Resolution & Audit Trail')
        ws_audit.append(['Ticket #', 'Resolution Notes', 'Resolved By', 'Resolution Date'])
        for c in ws_audit[1]:
            c.fill = header_fill
            c.font = header_font

        ticket_ids = [t.id for t in tickets]

        # Latest comment per ticket
        try:
            comments = (TicketComment.objects.filter(ticket_id__in=ticket_ids)
                        .select_related('user')
                        .order_by('ticket_id', '-created_at'))
            latest_comment = {}
            for c in comments:
                if c.ticket_id not in latest_comment:
                    latest_comment[c.ticket_id] = c
        except Exception:
            latest_comment = {}

        audit_logs = (AuditLog.objects.select_related('user', 'ticket')
                      .filter(ticket_id__in=ticket_ids)
                      .order_by('ticket_id', 'action_time'))
        audit_by_ticket = {}
        for a in audit_logs:
            audit_by_ticket.setdefault(a.ticket_id, []).append(a)

        for t in tickets:
            notes = '—'
            if t.id in latest_comment:
                c = latest_comment[t.id]
                notes = (c.comment or '')[:800] or '—'

            resolved_by = '—'
            res_date = t.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if t.resolved_at else '—'

            logs = audit_by_ticket.get(t.id, [])
            resolved_log = next((a for a in logs if a.action_type == 'ticket_resolved'), None)
            if resolved_log and resolved_log.user:
                resolved_by = resolved_log.user.get_full_name() or resolved_log.user.username
            elif t.status in (TicketStatus.RESOLVED, TicketStatus.CLOSED):
                if logs and logs[-1].user:
                    resolved_by = logs[-1].user.get_full_name() or logs[-1].user.username

            ws_audit.append([t.id, notes, resolved_by, res_date])

        ws_audit.append([])
        ws_audit.append(['Ticket #', 'User', 'Action Performed', 'Date and Time', 'Previous Status', 'New Status'])
        for c in ws_audit[ws_audit.max_row]:
            c.fill = header_fill
            c.font = header_font

        for t in tickets:
            logs = audit_by_ticket.get(t.id, [])
            for a in logs[:200]:
                prev_status = (a.metadata or {}).get('previous_status')
                new_status = None
                if a.action_type == 'ticket_resolved':
                    new_status = 'Resolved'
                if a.action_type == 'ticket_closed':
                    new_status = 'Closed'
                if a.action_type == 'ticket_reopened':
                    new_status = 'In Progress'

                ws_audit.append([
                    t.id,
                    (a.user.get_full_name() if a.user else 'System'),
                    a.action_type,
                    a.action_time.strftime('%Y-%m-%d %H:%M:%S'),
                    prev_status or '—',
                    new_status or '—',
                ])

        _autosize_columns(ws_audit)

        # Audit log export record
        AuditLog.objects.create(
            user=request.user,
            action_type='ticket_exported',
            description=f'Ticket incident audit report exported as Excel by {request.user.username}.',
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            metadata={'format': 'excel', 'ticket_count': queryset.count()},
        )

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="incident_audit_report.xlsx"'
        return response



def get_device_audit_report_context():
    devices = list(Device.objects.all())
    
    # Calculate incident statistics per device ID
    ticket_stats = Ticket.objects.values('device').annotate(
        total_incidents=Count('id'),
        active_incidents=Count('id', filter=Q(status__in=[TicketStatus.OPEN, TicketStatus.IN_PROGRESS]))
    )
    stats_map = {s['device']: s for s in ticket_stats if s['device']}

    today = timezone.now().date()
    overdue_days_threshold = 180

    def device_to_view_model(device, stats):
        last_inspected = device.last_inspected
        next_inspection = device.next_inspection

        inspection_overdue = False
        if last_inspected and last_inspected <= today - timezone.timedelta(days=overdue_days_threshold):
            inspection_overdue = True
        if next_inspection and next_inspection <= today:
            inspection_overdue = True

        status_label = device.get_status_display() if hasattr(device, 'get_status_display') else device.status
        condition_label = device.get_condition_display() if hasattr(device, 'get_condition_display') else device.condition

        return {
            'serial_number': device.serial_number,
            'name': device.name,
            'wing': device.wing,
            'floor': device.floor,
            'room_number': device.room_number,
            'status': device.status,
            'status_label': status_label,
            'condition': device.condition,
            'condition_label': condition_label,
            'last_inspected': device.last_inspected,
            'next_inspection': device.next_inspection,
            'inspection_due': inspection_overdue,
            'recommendation': device.recommendation(),
            'total_incidents': stats.get('total_incidents', 0),
            'active_incidents': stats.get('active_incidents', 0),
            'notes': device.notes,
        }

    working = []
    attention = []
    not_working = []

    for device in devices:
        stats = stats_map.get(device.pk, {})
        view_model = device_to_view_model(device, stats)
        if device.status == DeviceStatus.OPERATIONAL and device.condition == DeviceCondition.GOOD:
            working.append(view_model)
        elif device.status == DeviceStatus.DECOMMISSIONED or device.condition == DeviceCondition.POOR:
            not_working.append(view_model)
        elif device.needs_attention:
            attention.append(view_model)
        else:
            attention.append(view_model)

    return {
        'generated_at': timezone.now(),
        'working_count': len(working),
        'attention_count': len(attention),
        'not_working_count': len(not_working),
        'total_devices': len(devices),
        'sections': [
            {'title': 'Working (Operational & Good)', 'devices': sorted(working, key=lambda x: x['serial_number'])},
            {'title': 'Needs Attention', 'devices': sorted(attention, key=lambda x: x['serial_number'])},
            {'title': 'Not Working', 'devices': sorted(not_working, key=lambda x: x['serial_number'])},
        ],
    }


def format_device_audit_date(value):
    return value.strftime('%Y-%m-%d') if value else '-'


def format_device_audit_location(device):
    return ', '.join(
        part for part in [device.get('wing'), device.get('floor'), device.get('room_number')] if part
    ) or '-'


def log_device_audit_export(request, export_format, device_count):
    AuditLog.objects.create(
        user=request.user,
        action_type='device_audit_exported',
        description=f'Device audit report exported as {export_format.upper()} by {request.user.username}.',
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', ''),
        metadata={'format': export_format, 'device_count': device_count},
    )


class DeviceListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Device
    template_name = 'core/device_list.html'
    context_object_name = 'devices'
    login_url = reverse_lazy('login')

    def test_func(self):
        return self.request.user.is_admin or self.request.user.is_technician

    def handle_no_permission(self):
        return HttpResponse('Access denied', status=403)


class DeviceCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Device
    form_class = DeviceForm
    template_name = 'core/device_form.html'
    success_url = reverse_lazy('device_list')
    login_url = reverse_lazy('login')

    def test_func(self):
        return self.request.user.is_admin or self.request.user.is_technician

    def handle_no_permission(self):
        return HttpResponse('Access denied', status=403)

    def form_valid(self, form):
        response = super().form_valid(form)
        AuditLog.objects.create(
            user=self.request.user,
            action_type='device_created',
            description=f"New device '{self.object.name}' (S/N: {self.object.serial_number}) added by {self.request.user.username}.",
            ip_address=get_client_ip(self.request),
            user_agent=self.request.META.get('HTTP_USER_AGENT', ''),
        )
        return response


class DeviceDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Device
    template_name = 'core/device_detail.html'
    context_object_name = 'device'
    login_url = reverse_lazy('login')

    def test_func(self):
        return self.request.user.is_admin or self.request.user.is_technician

    def handle_no_permission(self):
        return HttpResponse('Access denied', status=403)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['related_tickets'] = self.object.tickets.all().order_by('-created_at')
        return context


class DeviceUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Device
    form_class = DeviceForm
    template_name = 'core/device_form.html'
    success_url = reverse_lazy('device_list')
    login_url = reverse_lazy('login')

    def test_func(self):
        return self.request.user.is_admin or self.request.user.is_technician

    def handle_no_permission(self):
        return HttpResponse('Access denied', status=403)

    def form_valid(self, form):
        changed_fields = form.changed_data
        response = super().form_valid(form)
        AuditLog.objects.create(
            user=self.request.user,
            action_type='device_updated',
            description=f"Device '{self.object.name}' (S/N: {self.object.serial_number}) updated by {self.request.user.username}. Changed fields: {', '.join(changed_fields)}",
            ip_address=get_client_ip(self.request),
            user_agent=self.request.META.get('HTTP_USER_AGENT', ''),
            metadata={'changed_fields': changed_fields},
        )
        return response


class DeviceAuditReportView(LoginRequiredMixin, TemplateView):
    template_name = 'core/device_audit_report.html'
    login_url = reverse_lazy('login')

    def get_context_data(self, **kwargs):
        context = kwargs.copy()
        context['view'] = self
        context.update(get_device_audit_report_context())
        return context


class DeviceAuditReportPDFView(LoginRequiredMixin, View):
    login_url = reverse_lazy('login')

    def get(self, request, *args, **kwargs):
        report = get_device_audit_report_context()
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            title='EAC Device Audit Report',
            leftMargin=0.35 * inch,
            rightMargin=0.35 * inch,
            topMargin=0.45 * inch,
            bottomMargin=0.45 * inch,
        )
        styles = getSampleStyleSheet()
        normal = styles['Normal']

        def _draw_page_number(canvas: Canvas, doc_obj):
            canvas.saveState()
            canvas.setFont('Helvetica', 8)
            canvas.drawRightString(10.5 * inch, 0.25 * inch, f"Page {doc_obj.page}")
            canvas.restoreState()

        def _cell(value):
            return Paragraph(escape(str(value or '-')), normal)

        elements = [
            Paragraph('EAC Helpdesk Device Audit Report', styles['Title']),
            Paragraph(f"Generated by: {request.user.get_full_name() or request.user.username}", normal),
            Paragraph(f"Generated at: {report['generated_at']:%Y-%m-%d %H:%M:%S}", normal),
            Spacer(1, 12),
        ]

        summary_table = Table([
            ['Working (Good)', report['working_count']],
            ['Needs Attention', report['attention_count']],
            ['Not Working', report['not_working_count']],
            ['Total Devices', report['total_devices']],
        ], colWidths=[2.2 * inch, 1.0 * inch], hAlign='LEFT')
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 12))

        headers = [
            'Serial Number', 'Name', 'Location', 'Status', 'Condition', 'Incidents (T/A)',
            'Last Inspected', 'Next Inspection', 'Due', 'Recommendation'
        ]
        col_widths = [
            1.0 * inch, 1.0 * inch, 1.1 * inch, 0.7 * inch, 0.7 * inch, 0.8 * inch,
            0.9 * inch, 0.9 * inch, 0.45 * inch, 2.8 * inch
        ]

        for section in report['sections']:
            elements.append(Paragraph(escape(section['title']), styles['Heading2']))
            if not section['devices']:
                elements.append(Paragraph('No devices found in this category.', normal))
                elements.append(Spacer(1, 10))
                continue

            table_data = [headers]
            for device in section['devices']:
                table_data.append([
                    _cell(device['serial_number']),
                    _cell(device['name']),
                    _cell(format_device_audit_location(device)),
                    _cell(device['status_label']),
                    _cell(device['condition_label']),
                    _cell(f"{device['total_incidents']} / {device['active_incidents']}"),
                    _cell(format_device_audit_date(device['last_inspected'])),
                    _cell(format_device_audit_date(device['next_inspection'])),
                    _cell('Yes' if device['inspection_due'] else 'No'),
                    _cell(device['recommendation']),
                ])

            table = Table(table_data, repeatRows=1, colWidths=col_widths, hAlign='LEFT')
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 12))

        doc.build(elements, onFirstPage=_draw_page_number, onLaterPages=_draw_page_number)
        log_device_audit_export(request, 'pdf', report['total_devices'])

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="device_audit_report.pdf"'
        return response


class DeviceAuditReportExcelView(LoginRequiredMixin, View):
    login_url = reverse_lazy('login')

    def get(self, request, *args, **kwargs):
        from openpyxl.styles import Alignment, Font, PatternFill

        report = get_device_audit_report_context()
        workbook = Workbook()
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        summary = workbook.active
        summary.title = 'Summary'
        summary.append(['EAC Helpdesk Device Audit Report'])
        summary.append([f"Generated by: {request.user.get_full_name() or request.user.username}"])
        summary.append([f"Generated at: {report['generated_at']:%Y-%m-%d %H:%M:%S}"])
        summary.append([])
        summary.append(['Metric', 'Count'])
        for cell in summary[5]:
            cell.fill = header_fill
            cell.font = header_font
        summary.append(['Working (Good)', report['working_count']])
        summary.append(['Needs Attention', report['attention_count']])
        summary.append(['Not Working', report['not_working_count']])
        summary.append(['Total Devices', report['total_devices']])
        summary.column_dimensions['A'].width = 28
        summary.column_dimensions['B'].width = 14

        details = workbook.create_sheet('Devices')
        details.append([
            'Category', 'Serial Number', 'Name', 'Wing', 'Floor', 'Room Number',
            'Status', 'Condition', 'Last Inspected', 'Next Inspection',
            'Inspection Due', 'Recommendation', 'Notes'
        ])
        for cell in details[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        for section in report['sections']:
            for device in section['devices']:
                details.append([
                    section['title'],
                    device['serial_number'],
                    device['name'] or '-',
                    device['wing'] or '-',
                    device['floor'] or '-',
                    device['room_number'] or '-',
                    device['status_label'],
                    device['condition_label'],
                    format_device_audit_date(device['last_inspected']),
                    format_device_audit_date(device['next_inspection']),
                    'Yes' if device['inspection_due'] else 'No',
                    device['recommendation'],
                    device['notes'] or '-',
                ])

        for column in details.columns:
            column_letter = column[0].column_letter
            max_length = max(len(str(cell.value or '')) for cell in column)
            details.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 55)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        log_device_audit_export(request, 'excel', report['total_devices'])

        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="device_audit_report.xlsx"'
        return response


class DeviceAuditReportCSVView(LoginRequiredMixin, View):
    login_url = reverse_lazy('login')

    def get(self, request, *args, **kwargs):
        report = get_device_audit_report_context()
        csv_buffer = StringIO()
        writer = csv.writer(csv_buffer)
        writer.writerow([
            'Category', 'Serial Number', 'Name', 'Wing', 'Floor', 'Room Number',
            'Status', 'Condition', 'Last Inspected', 'Next Inspection',
            'Inspection Due', 'Recommendation', 'Notes'
        ])

        for section in report['sections']:
            for device in section['devices']:
                writer.writerow([
                    section['title'],
                    device['serial_number'],
                    device['name'] or '-',
                    device['wing'] or '-',
                    device['floor'] or '-',
                    device['room_number'] or '-',
                    device['status_label'],
                    device['condition_label'],
                    format_device_audit_date(device['last_inspected']),
                    format_device_audit_date(device['next_inspection']),
                    'Yes' if device['inspection_due'] else 'No',
                    device['recommendation'],
                    device['notes'] or '-',
                ])

        log_device_audit_export(request, 'csv', report['total_devices'])
        response = HttpResponse(csv_buffer.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="device_audit_report.csv"'
        return response


class TicketReportCSVView(LoginRequiredMixin, View):

    login_url = reverse_lazy('login')

    def dispatch(self, request, *args, **kwargs):
        if not (request.user.is_admin or request.user.is_technician):
            return HttpResponse('Access denied', status=403)
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        queryset = apply_ticket_report_filters(get_report_ticket_queryset(request.user), request.GET)
        csv_buffer = StringIO()
        writer = csv.writer(csv_buffer)
        writer.writerow(['Title', 'Serial', 'Status', 'Priority', 'Reported By', 'Assigned To', 'Updated At', 'Category'])
        for ticket in queryset.order_by('-updated_at'):
            writer.writerow([
                ticket.title,
                ticket.device_serial_number or 'N/A',
                ticket.get_status_display(),
                ticket.get_priority_display(),
                ticket.reported_by.get_full_name() or ticket.reported_by.username,
                ticket.assigned_to.get_full_name() if ticket.assigned_to else 'Unassigned',
                ticket.updated_at.strftime('%Y-%m-%d %H:%M'),
                ticket.category or 'N/A',
            ])
        AuditLog.objects.create(
            user=request.user,
            action_type='ticket_exported',
            description=f'Ticket report exported as CSV by {request.user.username}.',
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            metadata={'format': 'csv', 'ticket_count': queryset.count()},
        )
        response = HttpResponse(csv_buffer.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="ticket_report.csv"'
        return response


class KnowledgeBaseListView(LoginRequiredMixin, ListView):
    model = KnowledgeBaseArticle
    template_name = 'core/kb_list.html'
    context_object_name = 'articles'

    def get_queryset(self):
        queryset = KnowledgeBaseArticle.objects.filter(is_published=True)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) | Q(content__icontains=search) | Q(category__icontains=search)
            )
        return queryset


class KnowledgeBaseDetailView(LoginRequiredMixin, DetailView):
    model = KnowledgeBaseArticle
    template_name = 'core/kb_detail.html'
    context_object_name = 'article'


class KnowledgeBaseCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = KnowledgeBaseArticle
    form_class = KnowledgeBaseForm
    template_name = 'core/kb_form.html'
    success_url = reverse_lazy('kb_list')

    def test_func(self):
        return self.request.user.is_admin or self.request.user.is_technician

    def form_valid(self, form):
        form.instance.author = self.request.user
        return super().form_valid(form)


class MaterialRequestCreateView(LoginRequiredMixin, CreateView):
    model = MaterialRequest
    form_class = MaterialRequestForm
    template_name = 'core/material_request_form.html'
    success_url = reverse_lazy('material_request_list')
    login_url = reverse_lazy('login')

    def form_valid(self, form):
        form.instance.requester = self.request.user
        response = super().form_valid(form)
        AuditLog.objects.create(
            user=self.request.user,
            action_type='material_request_created',
            description=f"Material request for {self.object.quantity}x {self.object.material.name} created by {self.request.user.username}.",
            ip_address=get_client_ip(self.request),
            user_agent=self.request.META.get('HTTP_USER_AGENT', ''),
            metadata={'request_id': self.object.pk, 'material_id': self.object.material.pk, 'quantity': self.object.quantity},
        )
        messages.success(self.request, "Your material request has been submitted.")
        return response


class MaterialRequestListView(LoginRequiredMixin, ListView):
    model = MaterialRequest
    template_name = 'core/material_request_list.html'
    context_object_name = 'requests'
    paginate_by = 10
    login_url = reverse_lazy('login')

    def get_queryset(self):
        user = self.request.user
        queryset = MaterialRequest.objects.select_related('material', 'requester', 'approver').all()

        if user.is_property_manager or user.is_admin:
            # Property managers and admins see all requests
            pass
        else:
            # Staff only see their own requests
            queryset = queryset.filter(requester=user)

        status_filter = self.request.GET.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        material_filter = self.request.GET.get('material')
        if material_filter:
            queryset = queryset.filter(material__name__icontains=material_filter)

        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = MaterialRequestStatus.choices
        context['current_status_filter'] = self.request.GET.get('status', '')
        context['current_material_filter'] = self.request.GET.get('material', '')
        return context


class MaterialRequestDetailView(LoginRequiredMixin, DetailView):
    model = MaterialRequest
    template_name = 'core/material_request_detail.html'
    context_object_name = 'request'
    login_url = reverse_lazy('login')

    def get_queryset(self):
        user = self.request.user
        queryset = MaterialRequest.objects.select_related('material', 'requester', 'approver').all()
        if user.is_property_manager or user.is_admin:
            return queryset
        return queryset.filter(requester=user)

    def dispatch(self, request, *args, **kwargs):
        obj = self.get_object()
        if not (request.user.is_property_manager or request.user.is_admin or obj.requester == request.user):
            return HttpResponse('Access denied', status=403)
        return super().dispatch(request, *args, **kwargs)


class MaterialRequestActionView(LoginRequiredMixin, UserPassesTestMixin, View):
    login_url = reverse_lazy('login')
    action_type = None # 'approve', 'reject', 'issue'
    new_status = None
    success_message = ""
    error_message = ""

    def test_func(self):
        return self.request.user.is_property_manager or self.request.user.is_admin

    def handle_no_permission(self):
        messages.error(self.request, "You do not have permission to perform this action.")
        return redirect(reverse_lazy('material_request_list'))

    def post(self, request, pk, *args, **kwargs):
        material_request = get_object_or_404(MaterialRequest, pk=pk)

        if material_request.status == self.new_status:
            messages.info(request, f"Request is already {self.new_status}.")
            return redirect('material_request_detail', pk=pk)

        if self.new_status == MaterialRequestStatus.ISSUED and material_request.material.stock_quantity < material_request.quantity:
            messages.error(request, f"Cannot issue: Insufficient stock for {material_request.material.name}. Available: {material_request.material.stock_quantity}, Requested: {material_request.quantity}.")
            return redirect('material_request_detail', pk=pk)

        old_status = material_request.status
        material_request.status = self.new_status
        material_request.approver = request.user
        material_request.save()

        AuditLog.objects.create(
            user=request.user,
            action_type=f'material_request_{self.action_type}',
            description=f"Material request for {material_request.material.name} (ID: {material_request.pk}) {self.action_type}d by {request.user.username}.",
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            metadata={
                'request_id': material_request.pk,
                'old_status': old_status,
                'new_status': self.new_status,
                'approver_id': request.user.pk
            },
        )
        messages.success(request, self.success_message)
        return redirect('material_request_detail', pk=pk)


class MaterialRequestApproveView(MaterialRequestActionView):
    action_type = 'approved'
    new_status = MaterialRequestStatus.APPROVED
    success_message = "Material request approved successfully."


class MaterialRequestRejectView(MaterialRequestActionView):
    action_type = 'rejected'
    new_status = MaterialRequestStatus.REJECTED
    success_message = "Material request rejected."


class MaterialRequestIssueView(MaterialRequestActionView):
    action_type = 'issued'
    new_status = MaterialRequestStatus.ISSUED
    success_message = "Material request issued and stock deducted."


class KnowledgeBaseUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = KnowledgeBaseArticle
    form_class = KnowledgeBaseForm
    template_name = 'core/kb_form.html'
    success_url = reverse_lazy('kb_list')

    def test_func(self):
        return self.request.user.is_admin or self.request.user.is_technician
